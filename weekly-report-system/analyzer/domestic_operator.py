"""
分析模块 - 国内运营商
基于合并后的数据生成分析报告
"""

import os
import yaml
from datetime import datetime


class DomesticOperatorAnalyzer:
    """国内运营商模块分析器"""

    def __init__(self, config_path="config.yaml"):
        with open(config_path, "r", encoding="utf-8") as f:
            self.config = yaml.safe_load(f)

    def analyze(self, module_id, input_path=None):
        """
        分析周报数据
        参数:
            module_id: 模块 ID
            input_path: 合并后的 Excel 路径（可选）
        返回:
            dict: 分析结果
        """
        # 默认输入路径
        if input_path is None:
            from openpyxl import load_workbook
            import glob as g

            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            output_dir = os.path.join(project_root, "output")
            module_name = self._get_module_name(module_id)
            pattern = os.path.join(output_dir, f"【{module_name}】周报汇总_*.xlsx")
            files = g.glob(pattern)
            if not files:
                return {"error": f"未找到合并文件: {pattern}"}
            input_path = sorted(files)[-1]  # 最新的

        if not os.path.exists(input_path):
            return {"error": f"文件不存在: {input_path}"}

        from openpyxl import load_workbook
        wb = load_workbook(input_path, data_only=True)

        result = self._analyze_workbook(wb, module_id)
        wb.close()
        return result

    def _get_module_name(self, module_id):
        modules = self.config.get("modules", [])
        for m in modules:
            if m.get("id") == module_id:
                return m.get("name", module_id)
        return module_id

    def _analyze_workbook(self, wb, module_id):
        """分析工作簿"""
        # 跳过目录和分析 Sheet
        skip_sheets = {"📑 目录", "📊 本周分析"}

        people = []
        total_rows = 0
        risks = []
        projects = []

        for sheet_name in wb.sheetnames:
            if sheet_name in skip_sheets:
                continue

            ws = wb[sheet_name]
            person_data = {
                "name": sheet_name,
                "row_count": 0,
                "items": [],
            }

            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                cleaned = [str(c).strip() if c is not None else "" for c in row]
                person_data["row_count"] += 1
                person_data["items"].append(cleaned)

                # 提取风险 (F列)
                if len(cleaned) >= 6 and cleaned[5]:
                    risks.append({"person": sheet_name, "content": cleaned[5]})

                # 提取重点项目 (B列)
                if len(cleaned) >= 2 and cleaned[1]:
                    projects.append({"person": sheet_name, "content": cleaned[1]})

            people.append(person_data)
            total_rows += person_data["row_count"]

        # 构建分析结果
        result = {
            "module": module_id,
            "analyzed_at": datetime.now().isoformat(),
            "total_people": len(people),
            "total_items": total_rows,
            "risks": risks,
            "risk_count": len(risks),
            "projects": projects,
            "project_count": len(projects),
            "workload": sorted(
                [{"name": p["name"], "rows": p["row_count"]} for p in people],
                key=lambda x: x["rows"],
                reverse=True,
            ),
        }

        return result

    def format_report(self, result):
        """将分析结果格式化为可读文本"""
        if "error" in result:
            return f"❌ {result['error']}"

        lines = [
            "=" * 40,
            f"  国内运营商 - 本周分析报告",
            "=" * 40,
            "",
            f"📋 总览: {result['total_people']} 人提交, "
            f"{result['total_items']} 条工作, "
            f"{result['risk_count']} 项风险",
            "",
        ]

        if result["risks"]:
            lines.append("⚠️ 风险及求助:")
            for r in result["risks"]:
                lines.append(f"  • [{r['person']}] {r['content'][:80]}")
            lines.append("")

        lines.append("📊 工作量 TOP5:")
        for w in result["workload"][:5]:
            bar = "█" * min(w["rows"], 20)
            lines.append(f"  {w['name']:8s} {bar} {w['rows']} 条")

        return "\n".join(lines)
