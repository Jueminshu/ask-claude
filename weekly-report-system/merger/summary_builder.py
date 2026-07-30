"""
通用汇总层
输入：4 个模块的标准化摘要 dict（get_summary() 返回值）
输出：四模块总汇总 Excel（含执行摘要首页、模块一览表、风险关注、重点事项）

供下游使用：build_total_summary() 返回文件路径
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class SummaryBuilder:
    """通用汇总构建器

    汇集 4 个模块合并器的标准化摘要，生成一份「四模块总汇总」Excel 文件。
    与上游 ExcelMerger.get_summary() 返回的 dict 直接对接。
    """

    # 样式常量
    TITLE_FONT = Font(name="微软雅黑", size=16, bold=True)
    SUBTITLE_FONT = Font(name="微软雅黑", size=11)
    SECTION_FONT = Font(name="微软雅黑", size=13, bold=True)
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    WARN_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    GRAY_FONT = Font(color="999999")

    def __init__(self, output_dir="output"):
        """初始化构建器

        参数:
            output_dir: 输出目录（存放生成的汇总 Excel）
        """
        self.output_dir = output_dir

    # ---- 公共接口 ----

    def build_total_summary(self, summaries, week_start, week_end):
        """构建四模块总汇总 Excel

        参数:
            summaries: list[dict] — 4 个模块合并器的 get_summary() 返回值
                       每个 dict 应包含:
                         - module_name      (str)  模块名称
                         - total_people     (int)  应提交人数
                         - submitted_count  (int)  已提交人数
                         - submission_rate  (str)  提交率，如 "90%"
                         - risk_items       (list) 风险项字符串列表
                         - key_projects     (list) 重点项目字符串列表
                         - output_file      (str)  模块合并输出文件路径（可选）
                         - deadline_passed  (bool) 是否已过截止时间（可选）
            week_start: str   周期起始日期，如 "2026-07-27"
            week_end:   str   周期结束日期，如 "2026-08-02"

        返回:
            str  输出文件路径；若 summaries 为空则返回 None
        """
        if not summaries:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "执行摘要"

        # 写入各个区域
        self._write_executive_header(ws, week_start, week_end)
        current_row = self._write_module_table(ws, summaries)
        current_row = self._write_deadline_warning(ws, summaries, current_row)
        current_row = self._write_risk_section(ws, summaries, current_row)
        self._write_key_projects_section(ws, summaries, current_row)

        # 调整列宽
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14

        # 保存
        os.makedirs(self.output_dir, exist_ok=True)
        output_path = os.path.join(
            self.output_dir, f"四模块周报总汇总_{week_end}.xlsx"
        )
        wb.save(output_path)
        return output_path

    # ---- 内部写入方法 ----

    def _write_executive_header(self, ws, week_start, week_end):
        """写执行摘要标题和周期信息"""
        ws.merge_cells("A1:F1")
        title_cell = ws.cell(row=1, column=1, value="营销运作部 周报总汇总")
        title_cell.font = self.TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A2:F2")
        subtitle_cell = ws.cell(
            row=2, column=1, value=f"周期: {week_start} ~ {week_end}"
        )
        subtitle_cell.font = self.SUBTITLE_FONT
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

    def _write_module_table(self, ws, summaries):
        """写四模块提交一览表，返回下一可用行号"""
        # 表头
        headers = ["模块", "应提交", "已提交", "提交率", "风险数", "输出文件"]
        header_row = 4
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=ci, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.THIN_BORDER
            cell.alignment = self.CELL_ALIGNMENT

        # 数据行
        for ri, s in enumerate(summaries):
            row_idx = header_row + 1 + ri
            # 模块名称
            ws.cell(row=row_idx, column=1, value=s["module_name"]).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=1).alignment = self.CELL_ALIGNMENT
            # 应提交
            ws.cell(row=row_idx, column=2, value=s["total_people"]).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=2).alignment = self.CELL_ALIGNMENT
            # 已提交
            ws.cell(row=row_idx, column=3, value=s["submitted_count"]).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=3).alignment = self.CELL_ALIGNMENT
            # 提交率（低于 80% 标红）
            rate_cell = ws.cell(row=row_idx, column=4, value=s["submission_rate"])
            rate_cell.border = self.THIN_BORDER
            rate_cell.alignment = self.CELL_ALIGNMENT
            rate_num = int(s["submission_rate"].replace("%", ""))
            if rate_num < 80:
                rate_cell.fill = self.WARN_FILL
            # 风险数
            ws.cell(row=row_idx, column=5, value=len(s["risk_items"])).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=5).alignment = self.CELL_ALIGNMENT
            # 输出文件
            output_file = s.get("output_file", "")
            ws.cell(row=row_idx, column=6, value=output_file).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=6).alignment = self.CELL_ALIGNMENT

        # 合计行
        total_row = header_row + 1 + len(summaries)
        total_people = sum(s["total_people"] for s in summaries)
        total_submitted = sum(s["submitted_count"] for s in summaries)
        total_rate = f"{round(total_submitted / max(total_people, 1) * 100)}%"
        total_risks = sum(len(s["risk_items"]) for s in summaries)

        for ci in range(1, 7):
            cell = ws.cell(row=total_row, column=ci)
            cell.border = self.THIN_BORDER
            cell.alignment = self.CELL_ALIGNMENT

        ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=total_people)
        ws.cell(row=total_row, column=3, value=total_submitted)
        total_rate_cell = ws.cell(row=total_row, column=4, value=total_rate)
        if total_submitted / max(total_people, 1) < 0.8:
            total_rate_cell.fill = self.WARN_FILL
        ws.cell(row=total_row, column=5, value=total_risks)

        return total_row + 2  # 空一行

    def _write_deadline_warning(self, ws, summaries, start_row):
        """写截止时间预警（如有模块已过截止时间），返回下一可用行号"""
        passed_modules = [s for s in summaries if s.get("deadline_passed", False)]
        if not passed_modules:
            return start_row  # 无需预警

        ws.merge_cells(f"A{start_row}:F{start_row}")
        cell = ws.cell(
            row=start_row, column=1,
            value="⚠️ 以下模块已过提交截止时间，请注意跟进："
        )
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="CC0000")

        for i, s in enumerate(passed_modules):
            row_idx = start_row + 1 + i
            ws.cell(
                row=row_idx, column=1,
                value=f"    {s['module_name']}（提交率 {s['submission_rate']}）"
            ).font = Font(color="CC0000")

        return start_row + 1 + len(passed_modules) + 1  # 空一行

    def _write_risk_section(self, ws, summaries, start_row):
        """汇集所有模块的风险项，返回下一可用行号"""
        ws.merge_cells(f"A{start_row}:F{start_row}")
        ws.cell(row=start_row, column=1, value="⚠️ 风险关注").font = self.SECTION_FONT

        # 按模块收集风险项
        all_risks = []
        for s in summaries:
            for r in s["risk_items"]:
                all_risks.append(f"[{s['module_name']}] {r}")

        if all_risks:
            for i, risk in enumerate(all_risks):
                row_idx = start_row + 1 + i
                ws.merge_cells(f"A{row_idx}:F{row_idx}")
                ws.cell(
                    row=row_idx, column=1, value=f"{i + 1}. {risk}"
                ).alignment = self.LEFT_ALIGNMENT
            next_row = start_row + 1 + len(all_risks) + 1
        else:
            ws.merge_cells(f"A{start_row + 1}:F{start_row + 1}")
            ws.cell(
                row=start_row + 1, column=1, value="本周无风险项"
            ).font = self.GRAY_FONT
            next_row = start_row + 3

        return next_row

    def _write_key_projects_section(self, ws, summaries, start_row):
        """汇集所有模块的重点项目/事项"""
        ws.merge_cells(f"A{start_row}:F{start_row}")
        ws.cell(
            row=start_row, column=1, value="本周重点事项"
        ).font = self.SECTION_FONT

        # 按模块收集重点项目
        all_projects = []
        for s in summaries:
            for p in s["key_projects"]:
                all_projects.append(f"[{s['module_name']}] {p}")

        if all_projects:
            for i, proj in enumerate(all_projects):
                row_idx = start_row + 1 + i
                ws.merge_cells(f"A{row_idx}:F{row_idx}")
                ws.cell(
                    row=row_idx, column=1, value=f"{i + 1}. {proj}"
                ).alignment = self.LEFT_ALIGNMENT
        else:
            ws.merge_cells(f"A{start_row + 1}:F{start_row + 1}")
            ws.cell(
                row=start_row + 1, column=1, value="暂无记录"
            ).font = self.GRAY_FONT
