"""
合并模块
将多个人员的 Excel 周报合并为一个文件
包含：目录 Sheet（超链接）+ 分析 Sheet + 各人员数据 Sheet

配置驱动：列映射、表头行、数据起始行从 config.yaml 的 module_columns 读取
"""

import os
import copy
import yaml
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink


# 模块 ID 映射（DB 数字 ID → config 字符串 ID）
_MODULE_ID_TO_KEY = {
    1: "domestic_operator",
    2: "marketing_ops",
    3: "sales",
    4: "overseas_bd",
}
_MODULE_KEY_TO_ID = {v: k for k, v in _MODULE_ID_TO_KEY.items()}


class ExcelMerger:
    """Excel 合并器（配置驱动）"""

    # 样式
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
    LINK_FONT = Font(name="微软雅黑", size=11, color="0F766E", underline="single")
    TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def __init__(self, module_id, config_path="config.yaml"):
        """
        初始化合并器，从 config.yaml 读取列配置

        参数:
            module_id: 模块 ID（支持 DB 数字 ID 或 config 字符串 ID）
            config_path: 配置文件路径
        """
        # 解析模块 ID（支持数字 ID 和字符串 ID）
        if isinstance(module_id, int):
            self.module_db_id = module_id
            self.module_id = _MODULE_ID_TO_KEY.get(module_id, str(module_id))
        elif isinstance(module_id, str) and module_id.isdigit():
            self.module_db_id = int(module_id)
            self.module_id = _MODULE_ID_TO_KEY.get(self.module_db_id, module_id)
        else:
            self.module_id = str(module_id)
            self.module_db_id = _MODULE_KEY_TO_ID.get(self.module_id)

        # 读取配置文件
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        config_full_path = os.path.join(project_root, config_path)
        if not os.path.exists(config_full_path):
            config_full_path = config_path  # fallback: 当前工作目录
        with open(config_full_path, "r", encoding="utf-8") as f:
            self.config = yaml.safe_load(f)

        # 获取模块名称
        self.module_name = self._get_module_name(self.module_id)

        # 从 module_columns 读取列配置
        all_columns = self.config.get("module_columns", {})
        module_config = all_columns.get(self.module_id, {})
        self.columns = module_config.get("columns", {})
        self.data_start_row = module_config.get("data_start_row", 3)
        self.header_row = module_config.get("header_row", 2)
        self.total_people = 0  # 由 merge 流程设置

        if not self.columns:
            raise ValueError(f"模块 {self.module_id} 未配置列映射")

    @classmethod
    def resolve_module_id(cls, module_id):
        """
        将 DB 数字 ID 转换为 config 字符串 ID

        参数:
            module_id: int 或 str 类型的模块 ID
        返回:
            config 中使用的模块 ID 字符串
        """
        if isinstance(module_id, int):
            return _MODULE_ID_TO_KEY.get(module_id, str(module_id))
        if isinstance(module_id, str) and module_id.isdigit():
            return _MODULE_ID_TO_KEY.get(int(module_id), module_id)
        return str(module_id)

    def merge(self) -> dict:
        """
        合并指定模块的所有周报（本地文件模式），返回摘要 dict

        返回:
            dict: {
                module_name, total_people, submitted_count,
                submission_rate, risk_items, key_projects,
                output_file, deadline_passed
            }
        """
        local_config = self.config.get("local", {})
        scan_dir = local_config.get("scan_dir", f"data/raw/{self.module_id}")

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        scan_dir = os.path.join(project_root, scan_dir)
        output_dir = os.path.join(project_root, "output")
        os.makedirs(output_dir, exist_ok=True)

        # 扫描文件
        import glob
        pattern = os.path.join(scan_dir, "*.xlsx")
        files = [f for f in glob.glob(pattern) if not os.path.basename(f).startswith("~$")]

        if not files:
            print(f"[提示] 在 {scan_dir} 中未找到 Excel 文件")
            return {
                "module_name": self.module_name,
                "total_people": 0,
                "submitted_count": 0,
                "submission_rate": "0%",
                "risk_items": [],
                "key_projects": [],
                "output_file": None,
                "deadline_passed": False,
            }

        print(f"找到 {len(files)} 个文件，开始合并...")

        today = datetime.now().strftime("%Y%m%d")
        output_path = os.path.join(output_dir, f"【{self.module_name}】周报汇总_{today}.xlsx")

        # 收集所有人员数据
        all_data = []
        for filepath in files:
            filename = os.path.basename(filepath)
            person_name = self._extract_name(filename)
            person_data = self._extract_data(filepath, person_name)
            all_data.append(person_data)

        # 按姓名排序
        all_data.sort(key=lambda x: x["name"])

        # 构建合并工作簿
        self._build_merged_workbook(all_data, output_path, self.module_name, today)

        # 构建摘要
        summary = self._build_summary(output_path, all_data)
        return summary

    def merge_from_uploads(self, module_id, week_start, week_end):
        """
        从 Web 上传目录合并周报（供 app.py 调用）

        参数:
            module_id: DB 数字模块 ID（用于定位上传目录和 DB 记录）
            week_start: 周起始日期
            week_end: 周结束日期
        返回:
            输出文件路径，失败返回 None
        """
        import glob

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        upload_dir = os.path.join(project_root, "data", "uploads", week_start, str(module_id))

        if not os.path.exists(upload_dir):
            print(f"[提示] 上传目录不存在: {upload_dir}")
            return None

        pattern = os.path.join(upload_dir, "*.xlsx")
        files = [f for f in glob.glob(pattern) if not os.path.basename(f).startswith("~$")]

        if not files:
            print(f"[提示] 未找到上传文件")
            return None

        # 使用 self.module_id 对应的配置名称
        module_name = self._get_module_name_by_id(module_id) if module_id != self.module_db_id else self.module_name
        if module_id == self.module_db_id:
            module_name = self.module_name
        else:
            module_name = self._get_module_name_by_id(module_id)

        output_dir = os.path.join(project_root, "output")
        os.makedirs(output_dir, exist_ok=True)

        today = datetime.now().strftime("%Y%m%d")
        output_path = os.path.join(output_dir, f"【{module_name}】周报汇总_{today}.xlsx")

        all_data = []
        for filepath in files:
            filename = os.path.basename(filepath)
            person_name = self._extract_name(filename)
            person_data = self._extract_data(filepath, person_name)
            all_data.append(person_data)

        all_data.sort(key=lambda x: x["name"])

        # 更新提交记录的 row_count
        self._update_submission_counts(all_data, week_start, module_id)

        # 构建合并工作簿
        self._build_merged_workbook(all_data, output_path, module_name, today)

        return output_path

    def _build_merged_workbook(self, all_data, output_path, module_name, today):
        """构建合并工作簿（共享的合并逻辑）"""
        wb = Workbook()
        wb.remove(wb.active)

        # 第一步：创建目录 Sheet
        toc_sheet = wb.create_sheet("📑 目录", 0)
        self._build_toc(toc_sheet, all_data, module_name, today)

        # 第二步：创建分析 Sheet
        analysis_sheet = wb.create_sheet("📊 本周分析")
        self._build_analysis(analysis_sheet, all_data, module_name)

        # 第三步：为每个人创建 Sheet
        for person in all_data:
            sheet_name = person["name"][:31]
            ws = wb.create_sheet(sheet_name)
            self._build_person_sheet(ws, person)

        # 保存
        wb.save(output_path)
        self.total_people = len(all_data)
        print(f"合并完成: 共 {len(all_data)} 人，输出 {output_path}")

    def _build_summary(self, output_path, all_data=None) -> dict:
        """从合并文件提取摘要数据"""
        if all_data is None:
            all_data = []

        # 收集已提交人员名
        submitted_names = [p["name"] for p in all_data]
        total = self.total_people if hasattr(self, 'total_people') and self.total_people else len(submitted_names)

        # 收集风险项（从人员数据中提取）
        risk_items = []
        risk_col = self._get_risk_column()
        for person in all_data:
            for row in person.get("rows", []):
                if len(row) > risk_col:
                    risk_text = str(row[risk_col]).strip() if row[risk_col] else ""
                    if risk_text and risk_text != "None":
                        risk_items.append(f"{person['name']}: {risk_text[:80]}")

        # 收集重点项目
        key_projects = []
        for person in all_data:
            for row in person.get("rows", []):
                if len(row) >= 2 and row[1]:
                    proj_text = str(row[1]).strip()
                    if proj_text:
                        key_projects.append(f"{person['name']}: {proj_text}")

        submitted_count = len(submitted_names)
        return {
            "module_name": self.module_name,
            "total_people": total,
            "submitted_count": submitted_count,
            "submission_rate": f"{round(submitted_count / max(total, 1) * 100)}%",
            "risk_items": risk_items,
            "key_projects": key_projects,
            "output_file": output_path,
            "deadline_passed": False,
        }

    def _get_risk_column(self) -> int:
        """获取风险列的 0-based 索引（从 module_columns 配置读取）"""
        risk_col_letter = self.config.get("module_columns", {}).get(
            self.module_id, {}
        ).get("risk_column", "F")
        return ord(risk_col_letter.upper()) - ord('A')

    def _get_module_name_by_id(self, module_id):
        """根据数字 ID 从数据库获取模块名"""
        try:
            import sqlite3
            db_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "..", "data")
            db_path = os.path.join(db_dir, "weekly_report.db")
            if os.path.exists(db_path):
                conn = sqlite3.connect(db_path)
                row = conn.execute("SELECT name FROM modules WHERE id = ?", (module_id,)).fetchone()
                conn.close()
                if row:
                    return row[0]
        except Exception:
            pass
        return f"模块{module_id}"

    def _update_submission_counts(self, all_data, week_start, module_id):
        """更新提交记录的 row_count"""
        try:
            import sqlite3
            db_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "..", "data")
            db_path = os.path.join(db_dir, "weekly_report.db")
            if os.path.exists(db_path):
                conn = sqlite3.connect(db_path)
                for person in all_data:
                    conn.execute(
                        """UPDATE submissions SET row_count = ?
                           WHERE module_id = ? AND week_start = ?
                           AND file_path LIKE ?""",
                        (len(person["rows"]), module_id, week_start, f"%{person['name']}%")
                    )
                conn.commit()
                conn.close()
        except Exception:
            pass

    def _get_module_name(self, module_id):
        """
        获取模块显示名称
        优先从 config modules 列表查找，其次从 DB 查找
        """
        # 先从 config 查找
        modules = self.config.get("modules", [])
        for m in modules:
            if m.get("id") == module_id:
                return m.get("name", module_id)
        # 再从 DB 查找
        if self.module_db_id:
            try:
                return self._get_module_name_by_id(self.module_db_id)
            except Exception:
                pass
        return module_id

    def _extract_name(self, filename):
        """从文件名提取姓名"""
        # 去掉扩展名
        name = os.path.splitext(filename)[0]
        # 常见命名模式清理
        for prefix in ["周报_", "周报-", "weekly_", "weekly-"]:
            if name.lower().startswith(prefix):
                name = name[len(prefix):]
        for suffix in ["_周报", "-周报", "_weekly", "-weekly"]:
            if name.lower().endswith(suffix):
                name = name[: -len(suffix)]
        return name.strip()

    def _extract_data(self, filepath, person_name):
        """从 Excel 文件提取数据，跳过空行"""
        wb = load_workbook(filepath, data_only=True)
        rows = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 从数据起始行开始读取
            for row in ws.iter_rows(min_row=self.data_start_row, values_only=True):
                # 跳过完全空行
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                # 跳过仅有序号无实质内容的行（第2列起全空）
                meaningful = [c for c in row[1:] if c is not None and str(c).strip() != ""]
                if not meaningful:
                    continue
                cleaned = []
                for c in row:
                    if c is None:
                        cleaned.append("")
                    else:
                        cleaned.append(str(c).strip())
                rows.append(cleaned)

        wb.close()
        return {"name": person_name, "filepath": filepath, "rows": rows}

    def _build_toc(self, ws, all_data, module_name, date_str):
        """构建目录 Sheet"""
        ws.sheet_properties.tabColor = "0F766E"

        # 标题
        ws.merge_cells("A1:C1")
        ws["A1"] = f"【{module_name}】周报汇总目录"
        ws["A1"].font = self.TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # 日期
        ws.merge_cells("A2:C2")
        ws["A2"] = f"报告周期: {date_str}"
        ws["A2"].font = Font(name="微软雅黑", size=10, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")

        # 空行
        ws.append([])

        # 表头
        headers = ["序号", "人员", "数据行数"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # 数据行
        for idx, person in enumerate(all_data, 1):
            row = idx + 4
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=1).border = self.THIN_BORDER

            # 人名 + 超链接
            name_cell = ws.cell(row=row, column=2, value=person["name"])
            name_cell.font = self.LINK_FONT
            name_cell.border = self.THIN_BORDER
            # 创建内部超链接 (指向同名 Sheet)
            sheet_name = person["name"][:31]
            name_cell.hyperlink = Hyperlink(
                ref=name_cell.coordinate,
                location=f"'{sheet_name}'!A1",
                display=person["name"]
            )

            ws.cell(row=row, column=3, value=len(person["rows"])).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3).border = self.THIN_BORDER

        # 列宽
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 12

    def _build_analysis(self, ws, all_data, module_name):
        """构建分析 Sheet（配置驱动风险列索引）"""
        ws.sheet_properties.tabColor = "E67E22"

        total_people = len(all_data)
        total_rows = sum(len(p["rows"]) for p in all_data)

        # 使用配置驱动的风险列索引
        risk_col = self._get_risk_column()

        # 收集风险项和重点项目
        risks = []
        projects = []
        for person in all_data:
            for row in person["rows"]:
                if len(row) > risk_col and row[risk_col]:  # 风险列（配置驱动）
                    risks.append({"person": person["name"], "content": row[risk_col]})
                if len(row) >= 2 and row[1]:  # B列: 重点项目
                    projects.append({"person": person["name"], "content": row[1]})

        # --- 标题 ---
        ws.merge_cells("A1:E1")
        ws["A1"] = f"【{module_name}】本周分析"
        ws["A1"].font = self.TITLE_FONT

        # --- 总览 ---
        row = 3
        ws.cell(row=row, column=1, value="📋 本周总览").font = Font(name="微软雅黑", size=12, bold=True)
        row += 1
        ws.cell(row=row, column=1, value="提交人数").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_people)
        row += 1
        ws.cell(row=row, column=1, value="工作条目数").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_rows)
        row += 1
        ws.cell(row=row, column=1, value="风险/求助项").font = Font(bold=True)
        ws.cell(row=row, column=2, value=len(risks))

        # --- 风险汇总 ---
        row += 2
        ws.cell(row=row, column=1, value="⚠️ 风险及求助汇总").font = Font(name="微软雅黑", size=12, bold=True)
        row += 1
        risk_headers = ["人员", "风险/求助内容"]
        for col_idx, header in enumerate(risk_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        if risks:
            row += 1
            for r in risks:
                ws.cell(row=row, column=1, value=r["person"]).border = self.THIN_BORDER
                ws.cell(row=row, column=2, value=r["content"]).border = self.THIN_BORDER
                ws.cell(row=row, column=2).alignment = self.CELL_ALIGNMENT
                row += 1
        else:
            row += 1
            ws.cell(row=row, column=1, value="🎉 本周无风险/求助项")

        # --- 重点项目 ---
        row += 2
        ws.cell(row=row, column=1, value="🎯 重点项目一览").font = Font(name="微软雅黑", size=12, bold=True)
        row += 1
        proj_headers = ["人员", "重点项目"]
        for col_idx, header in enumerate(proj_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        if projects:
            row += 1
            for p in projects:
                ws.cell(row=row, column=1, value=p["person"]).border = self.THIN_BORDER
                ws.cell(row=row, column=2, value=p["content"]).border = self.THIN_BORDER
                ws.cell(row=row, column=2).alignment = self.CELL_ALIGNMENT
                row += 1

        # --- 工作量 ---
        row += 2
        ws.cell(row=row, column=1, value="📊 工作量概况").font = Font(name="微软雅黑", size=12, bold=True)
        row += 1
        work_headers = ["人员", "条目数"]
        for col_idx, header in enumerate(work_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        row += 1
        for person in sorted(all_data, key=lambda x: len(x["rows"]), reverse=True):
            ws.cell(row=row, column=1, value=person["name"]).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=len(person["rows"])).border = self.THIN_BORDER
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            row += 1

        # 列宽
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15

    def _build_person_sheet(self, ws, person):
        """构建个人数据 Sheet（配置驱动列映射）"""
        ws.sheet_properties.tabColor = "2E86C1"

        # 表头（从 self.columns 读取，而非硬编码 DOMESTIC_COLUMNS）
        headers = list(self.columns.values())
        num_cols = len(headers)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # 数据行（列数由配置决定）
        for row_idx, row_data in enumerate(person["rows"], 2):
            for col_idx, value in enumerate(row_data[:num_cols], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = self.CELL_ALIGNMENT
                cell.border = self.THIN_BORDER

        # 回到顶部 + 返回目录链接
        row_offset = len(person["rows"]) + 3
        ws.cell(row=row_offset, column=1, value="← 返回目录").font = self.LINK_FONT
        ws.cell(row=row_offset, column=1).hyperlink = Hyperlink(
            ref=ws.cell(row=row_offset, column=1).coordinate,
            location="'📑 目录'!A1",
            display="← 返回目录"
        )

        # 列宽（由配置列数决定，使用预定义宽度序列）
        default_widths = [6, 25, 30, 40, 40, 40]
        col_keys = list(self.columns.keys())
        for i, col_letter in enumerate(col_keys):
            width = default_widths[i] if i < len(default_widths) else 40
            ws.column_dimensions[col_letter].width = width
