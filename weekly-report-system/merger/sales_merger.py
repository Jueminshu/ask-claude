"""
销售部合并器
处理三段式 Excel 模板：日常工作 / 质量供应问题 / 竞品信息
每个 Section 独立合并为一张大表，所有人员数据纵向拼接。
"""

import os
import re
import yaml
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class SalesMerger:
    """销售部三段式 Excel 合并器

    与 ExcelMerger 的区别：
    - 不使用一人一 Sheet 模式，而是按 Section 横向展开
    - 每人每个 Section 的数据行纵向拼接在同一张总表中
    - 输出包含：目录 Sheet + 3 个 Section Sheet + 分析 Sheet
    """

    # 样式常量（与 excel_merger 保持一致的风格）
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
    TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self, config_path="config.yaml"):
        """
        初始化销售部合并器，从 config.yaml 读取三段式 Section 配置。

        参数:
            config_path: 配置文件路径（相对于项目根目录或当前工作目录）
        """
        # 读取配置文件（项目根目录优先）
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        config_full_path = os.path.join(project_root, config_path)
        if not os.path.exists(config_full_path):
            config_full_path = config_path  # fallback: 当前工作目录
        with open(config_full_path, "r", encoding="utf-8") as f:
            self.config = yaml.safe_load(f)

        # 从 module_columns.sales.sections 读取三段式配置
        sales_cfg = self.config.get("module_columns", {}).get("sales", {})
        self.sections = sales_cfg.get("sections", [])

        if not self.sections:
            raise ValueError("配置中未找到 module_columns.sales.sections，请检查 config.yaml")

    # ===== 主入口 =====

    def merge(self, upload_dir, output_dir, week_start, week_end):
        """
        扫描 upload_dir 中的所有 .xlsx 文件，按三段式分段合并。

        参数:
            upload_dir:  上传文件所在目录
            output_dir:  合并结果输出目录
            week_start:  周起始日期 (YYYY-MM-DD)
            week_end:    周结束日期 (YYYY-MM-DD)

        返回:
            标准化摘要 dict，包含:
            module_name, total_people, submitted_count, submission_rate,
            risk_items, key_projects, output_file, deadline_passed
        """
        # 扫描文件（过滤临时文件 ~$）
        files = [
            f for f in os.listdir(upload_dir)
            if f.endswith((".xlsx", ".xls")) and not f.startswith("~$")
        ]
        files.sort()

        if not files:
            return {
                "module_name": "销售部",
                "total_people": 0,
                "submitted_count": 0,
                "submission_rate": "0%",
                "risk_items": [],
                "key_projects": [],
                "output_file": None,
                "deadline_passed": False,
            }

        # 创建工作簿（删除默认 Sheet）
        wb = Workbook()
        wb.remove(wb.active)

        # 1. 目录 Sheet
        toc_sheet = wb.create_sheet("📑 目录", 0)
        self._write_toc(toc_sheet, files)

        # 2. 每个 Section 创建一张总表
        person_data = {}  # {section_name: [(person_name, data_rows), ...]}

        for section in self.sections:
            section_name = section["name"]
            num_columns = len(section["columns"])
            sheet = wb.create_sheet(section_name)
            self._write_section_header(sheet, section)
            person_data[section_name] = []

            # 数据行从第 2 行开始（第 1 行是表头）
            row_offset = 1

            for fname in files:
                person_name = self._extract_person_name(fname)
                file_path = os.path.join(upload_dir, fname)

                try:
                    src = load_workbook(file_path, data_only=True)
                    src_sheet = src[src.sheetnames[0]]
                except Exception as e:
                    print(f"[警告] 无法读取文件 {fname}: {e}")
                    continue

                data_rows = self._extract_section_rows(
                    src_sheet,
                    section["start_row"],
                    section["end_row"],
                    num_columns,
                )
                src.close()

                if data_rows:
                    person_data[section_name].append((person_name, data_rows))
                    for dr in data_rows:
                        row_offset += 1
                        # A 列：提交人姓名
                        name_cell = sheet.cell(row=row_offset, column=1, value=person_name)
                        name_cell.alignment = self.CELL_ALIGNMENT
                        name_cell.border = self.THIN_BORDER
                        # B 列起：Section 数据
                        for ci, val in enumerate(dr, start=2):
                            cell = sheet.cell(row=row_offset, column=ci, value=val)
                            cell.alignment = self.CELL_ALIGNMENT
                            cell.border = self.THIN_BORDER

            # 设置列宽
            self._set_section_column_widths(sheet, section)

        # 3. 分析 Sheet
        self._build_analysis_sheet(wb, files, person_data)

        # 4. 保存文件
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"销售部_周报汇总_{week_end}.xlsx")
        wb.save(output_path)
        print(f"销售部合并完成: 共 {len(files)} 个文件 → {output_path}")

        return self._build_summary(files, person_data, output_path, week_start, week_end)

    # ===== 文件名解析 =====

    def _extract_person_name(self, filename):
        """
        从文件名提取姓名。

        示例:
            '时间+营销Team1-week27-张三周报.xlsx' → '张三'
            '张三.xlsx' → '张三'
        """
        name = os.path.splitext(filename)[0]
        # 尝试匹配 "XXX周报" 模式，提取前面的中文名
        m = re.search(r'[周报weekly]*([一-鿿]{2,4})周报', name)
        if m:
            return m.group(1)
        # 回退：取 "-" 分隔的最后一段，去掉 "周报" 后缀
        parts = name.replace("周报", "").split("-")
        last = parts[-1].strip() if parts else name.replace("周报", "")
        return last if last else "未知"

    # ===== 数据提取 =====

    def _extract_section_rows(self, sheet, start_row, end_row, num_columns):
        """
        从 Sheet 中提取指定区块的数据行。

        参数:
            sheet:       源工作表
            start_row:   区块表头行号（1-based，Excel 行号）
            end_row:     区块结束行号（1-based）
            num_columns: Section 配置的列数（只读取这么多列）

        返回:
            list[list]: 数据行列表，每行为单元格值列表
                        已跳过表头行和全空行
        """
        rows = []
        for row_idx in range(start_row + 1, end_row + 1):
            # start_row 是表头，数据从 start_row+1 开始
            row_data = []
            all_empty = True
            for col_idx in range(1, num_columns + 1):
                val = sheet.cell(row=row_idx, column=col_idx).value
                if val is not None and str(val).strip():
                    all_empty = False
                row_data.append(val)
            if not all_empty:
                rows.append(row_data)
        return rows

    # ===== 目录 Sheet =====

    def _write_toc(self, sheet, files):
        """写目录 Sheet：标题 + 提交人员清单"""
        sheet.sheet_properties.tabColor = "2B579A"

        # 标题
        sheet.merge_cells("A1:C1")
        cell = sheet.cell(row=1, column=1, value="销售部周报汇总")
        cell.font = self.TITLE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # 日期
        today = datetime.now().strftime("%Y-%m-%d")
        sheet.merge_cells("A2:C2")
        cell = sheet.cell(row=2, column=1, value=f"生成日期: {today}")
        cell.font = Font(name="微软雅黑", size=10, color="666666")
        cell.alignment = Alignment(horizontal="center")

        # 空行
        sheet.append([])

        # 表头
        headers = ["序号", "提交人员", "文件名"]
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=4, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # 数据行
        for i, fname in enumerate(files):
            row_idx = 5 + i
            name = self._extract_person_name(fname)
            cell_a = sheet.cell(row=row_idx, column=1, value=i + 1)
            cell_a.alignment = Alignment(horizontal="center")
            cell_a.border = self.THIN_BORDER
            cell_b = sheet.cell(row=row_idx, column=2, value=name)
            cell_b.border = self.THIN_BORDER
            cell_c = sheet.cell(row=row_idx, column=3, value=fname)
            cell_c.border = self.THIN_BORDER

        # 列宽
        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 50

    # ===== Section 表头 =====

    def _write_section_header(self, sheet, section):
        """
        写 Section 表头。

        A 列固定为「提交人」，B 列起使用 Section 配置的列名。
        """
        cols = list(section["columns"].values())

        # A 列：提交人
        cell = sheet.cell(row=1, column=1, value="提交人")
        cell.font = self.HEADER_FONT
        cell.fill = self.HEADER_FILL
        cell.alignment = self.HEADER_ALIGNMENT
        cell.border = self.THIN_BORDER

        # B 列起：Section 列名
        for ci, col_name in enumerate(cols, start=2):
            cell = sheet.cell(row=1, column=ci, value=col_name)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # 冻结首行
        sheet.freeze_panes = "A2"

    def _set_section_column_widths(self, sheet, section):
        """根据 Section 配置设置列宽"""
        num_cols = len(section["columns"]) + 1  # +1 为「提交人」列
        # A 列（提交人）
        sheet.column_dimensions["A"].width = 12
        # 数据列
        for i in range(2, num_cols + 1):
            col_letter = get_column_letter(i)
            sheet.column_dimensions[col_letter].width = 30

    # ===== 分析 Sheet =====

    def _build_analysis_sheet(self, wb, files, person_data):
        """生成分析 Sheet：提交率 + 各 Section 数据量 + 风险项"""
        sheet = wb.create_sheet("📊 本周分析")
        sheet.sheet_properties.tabColor = "E67E22"

        # 获取模块配置的总人数
        modules_list = self.config.get("modules", [])
        total_people = 0
        for m in modules_list:
            if m.get("id") == "sales" or m.get("name") == "销售部":
                members = m.get("members", [])
                total_people = len(members) if members else 0
                break
        if total_people == 0:
            total_people = len(files)

        submitted = len(files)

        # --- 标题 ---
        sheet.merge_cells("A1:E1")
        sheet.cell(row=1, column=1, value="销售部本周分析").font = self.TITLE_FONT

        # --- 总览 ---
        row = 3
        sheet.cell(row=row, column=1, value="📋 提交总览").font = Font(name="微软雅黑", size=12, bold=True)
        row += 1
        sheet.cell(row=row, column=1, value="提交人数").font = Font(bold=True)
        sheet.cell(row=row, column=2, value=f"{submitted} / {total_people}")
        row += 1
        sheet.cell(row=row, column=1, value="提交率").font = Font(bold=True)
        rate = f"{round(submitted / max(total_people, 1) * 100)}%"
        sheet.cell(row=row, column=2, value=rate)

        # --- 各 Section 数据量 ---
        row += 2
        sheet.cell(row=row, column=1, value="📊 各 Section 数据量").font = Font(name="微软雅黑", size=12, bold=True)
        row += 1
        for ci, header in enumerate(["Section", "数据条目数", "涉及人数"], start=1):
            cell = sheet.cell(row=row, column=ci, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        for section_name, data in person_data.items():
            row += 1
            total_rows = sum(len(rows) for _, rows in data)
            unique_people = len(data)
            sheet.cell(row=row, column=1, value=section_name).border = self.THIN_BORDER
            sheet.cell(row=row, column=2, value=total_rows).border = self.THIN_BORDER
            sheet.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            sheet.cell(row=row, column=3, value=unique_people).border = self.THIN_BORDER
            sheet.cell(row=row, column=3).alignment = Alignment(horizontal="center")

        # --- 风险项汇总（Section 2: 质量供应问题） ---
        row += 2
        sheet.cell(row=row, column=1, value="⚠️ 质量供应问题汇总").font = Font(name="微软雅黑", size=12, bold=True)
        row += 1
        risk_headers = ["提交人", "客户", "问题描述", "客户诉求"]
        for ci, header in enumerate(risk_headers, 1):
            cell = sheet.cell(row=row, column=ci, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # Section 2 数据
        section2_data = person_data.get("质量供应问题", [])
        risk_count = 0
        if section2_data:
            for person_name, rows in section2_data:
                for rd in rows:
                    # rd: [客户, 平台, 问题描述, 客户诉求, 影响, 是否重复]
                    customer = str(rd[0]) if len(rd) > 0 and rd[0] else ""
                    problem = str(rd[2]) if len(rd) > 2 and rd[2] else ""
                    demand = str(rd[3]) if len(rd) > 3 and rd[3] else ""
                    if problem.strip() or customer.strip():
                        row += 1
                        sheet.cell(row=row, column=1, value=person_name).border = self.THIN_BORDER
                        sheet.cell(row=row, column=2, value=customer).border = self.THIN_BORDER
                        sheet.cell(row=row, column=2).alignment = self.CELL_ALIGNMENT
                        sheet.cell(row=row, column=3, value=problem).border = self.THIN_BORDER
                        sheet.cell(row=row, column=3).alignment = self.CELL_ALIGNMENT
                        sheet.cell(row=row, column=4, value=demand).border = self.THIN_BORDER
                        sheet.cell(row=row, column=4).alignment = self.CELL_ALIGNMENT
                        risk_count += 1

        if risk_count == 0:
            row += 1
            sheet.cell(row=row, column=1, value="🎉 本周无质量供应问题")

        # --- 未提交人员 ---
        row += 2
        sheet.cell(row=row, column=1, value="📝 已提交人员清单").font = Font(name="微软雅黑", size=12, bold=True)
        row += 1
        for i, fname in enumerate(files):
            sheet.cell(row=row, column=1, value=f"{i + 1}. {self._extract_person_name(fname)}")
            row += 1

        # 列宽
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 22
        sheet.column_dimensions["C"].width = 45
        sheet.column_dimensions["D"].width = 45
        sheet.column_dimensions["E"].width = 20

    # ===== 摘要生成 =====

    def _build_summary(self, files, person_data, output_path, week_start, week_end):
        """
        构建标准化摘要 dict，供 app.py 和通知模块使用。

        返回格式与 ExcelMerger._build_summary 一致。
        """
        # 总人数（优先从 config modules 获取）
        modules_list = self.config.get("modules", [])
        total = len(files)
        for m in modules_list:
            if m.get("id") == "sales" or m.get("name") == "销售部":
                members = m.get("members", [])
                total = len(members) if members else len(files)
                break

        submitted = len(files)

        # 收集风险项（Section 2: 质量供应问题）
        risk_items = []
        section2_data = person_data.get("质量供应问题", [])
        for person_name, rows in section2_data:
            for rd in rows:
                # rd: [客户, 平台, 问题描述, 客户诉求, 影响, 是否重复]
                customer = str(rd[0]).strip() if len(rd) > 0 and rd[0] else ""
                problem = str(rd[2]).strip() if len(rd) > 2 and rd[2] else ""
                if customer or problem:
                    desc = f"{customer}: {problem}" if customer else problem
                    risk_items.append(f"{person_name}: {desc[:80]}")

        # 收集重点项目（Section 1 B 列: 本周工作进展）
        key_projects = []
        section1_data = person_data.get("日常工作", [])
        for person_name, rows in section1_data:
            for rd in rows:
                progress = str(rd[1]).strip() if len(rd) > 1 and rd[1] else ""
                if progress:
                    key_projects.append(f"{person_name}: {progress[:80]}")

        # 检查截止时间
        deadline_passed = self._check_deadline()

        return {
            "module_name": "销售部",
            "total_people": total,
            "submitted_count": submitted,
            "submission_rate": f"{round(submitted / max(total, 1) * 100)}%",
            "risk_items": risk_items,
            "key_projects": key_projects,
            "output_file": output_path,
            "deadline_passed": deadline_passed,
        }

    def _check_deadline(self):
        """检查是否已过截止时间（周日 24:00）"""
        now = datetime.now()
        weekday = now.weekday()  # 0=周一, 6=周日
        if weekday == 6:  # 周日
            return False  # 今天还是周日，未截止
        # 周一及以后
        return True
