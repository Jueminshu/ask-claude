"""
风险提取器
- Excel 模板（国内运营商、营销运营部）：从固定风险列直接读取
- PPT 模板（海外 BD）：Claude API 自由文本识别
"""
import os
import json
from database_v2 import get_db, upsert_risk_items


# === 模板 → 风险列名映射 ===
TEMPLATE_RISK_COLUMNS = {
    "周报": ["风险及求助"],            # 国内运营商
    "周报_营销": ["问题反馈&所需支持"],  # 营销运营部
}


def _find_risk_column(headers, template_name):
    """根据模板名查找风险列索引"""
    candidates = TEMPLATE_RISK_COLUMNS.get(template_name, [])
    if not candidates:
        # 模糊匹配：任意 header 包含 "风险" 或 "问题"
        for i, h in enumerate(headers):
            h_str = str(h).lower() if h else ""
            if any(kw in h_str for kw in ["风险", "求助", "问题反馈"]):
                return i, str(h)
        return None, None
    for i, h in enumerate(headers):
        if str(h) in candidates:
            return i, str(h)
    return None, None


def _parse_excel_headers(file_path):
    """读取 Excel 文件所有数据 sheet 的 headers 和 rows。跳过目录/分析 sheet。"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        all_headers = []
        all_data_rows = []
        skip_keywords = ["目录", "分析", "汇总", "概述"]

        for name in wb.sheetnames:
            if any(kw in name for kw in skip_keywords):
                continue
            ws = wb[name]
            if ws.max_row < 2:
                continue
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and any(c for c in row):
                    all_data_rows.append([str(c) if c else "" for c in row])
            if not all_headers:
                all_headers = headers

        wb.close()
        return all_headers, all_data_rows
    except Exception:
        return [], []


def extract_risks_from_excel(file_path, template_name, user_id, module_id, week_start, submission_file_id):
    """
    从 Excel 文件的风险列直接提取风险条目。
    返回: list[dict]
    """
    headers, data_rows = _parse_excel_headers(file_path)
    if not headers:
        return []

    risk_col_idx, col_name = _find_risk_column(headers, template_name)
    if risk_col_idx is None:
        return []

    # 找到 "重点项目" 或 "业务模块" 列作为 customer 来源
    customer_col_idx = None
    for i, h in enumerate(headers):
        h_str = str(h) if h else ""
        if any(kw in h_str for kw in ["重点项目", "业务模块", "客户"]):
            customer_col_idx = i
            break

    risks = []
    for row in data_rows:
        if risk_col_idx >= len(row) or not row[risk_col_idx].strip():
            continue  # 风险列为空，跳过

        customer = row[customer_col_idx] if customer_col_idx is not None and customer_col_idx < len(row) else None
        risk_text = row[risk_col_idx].strip()

        # 基本严重程度判断
        severity = "medium"
        high_keywords = ["严重", "紧急", "高风险", "损失", "投诉升级"]
        low_keywords = ["已完成", "已解决", "轻微"]
        if any(kw in risk_text for kw in high_keywords):
            severity = "high"
        elif any(kw in risk_text for kw in low_keywords):
            severity = "low"

        risks.append({
            "customer": customer,
            "risk_description": risk_text,
            "severity": severity,
            "is_new": 1,
            "source_column": col_name,
        })

    return risks


def extract_risks_via_claude(file_path, file_type, module_id, week_start):
    """
    对 PPT 等非结构化文件，使用 Claude API 从文本中提取风险。
    返回: list[dict]
    """
    from services.file_parser import extract_text

    text = extract_text(file_path, file_type)
    if not text or text.startswith("["):
        return []  # 提取失败，返回空

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return _keyword_risk_fallback(text)

    import urllib.request
    import urllib.error

    prompt = f"""你是一个风险分析助手。从以下周报文本中提取所有风险条目。

周报文本：
{text[:6000]}

对每个风险条目，返回以下字段：
- customer: 关联的客户或项目名称（如无则填 null）
- risk_description: 风险描述（简要一句话）
- severity: 严重程度（high/medium/low）
- is_new: 是否本周新提出（1=新风险, 0=持续风险）

请以 JSON 数组格式返回，只返回 JSON，不要其他文字。
如果无风险，返回空数组 []。

示例返回格式：
[{{"customer": "政企客户", "risk_description": "竞品降价导致续约困难", "severity": "high", "is_new": 1}}]"""

    try:
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=json.dumps({
                "model": "claude-sonnet-5",
                "max_tokens": 1024,
                "messages": [{"role": "user", "content": prompt}],
            }).encode("utf-8"),
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
            },
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read())
        content = result["content"][0]["text"]
        json_start = content.find("[")
        json_end = content.rfind("]") + 1
        if json_start >= 0 and json_end > json_start:
            risks = json.loads(content[json_start:json_end])
            for r in risks:
                r.setdefault("source_column", "claude_api")
            return risks
        return []
    except Exception:
        return _keyword_risk_fallback(text)


def _keyword_risk_fallback(text):
    """关键词回退：从文本中识别风险信号"""
    risks = []
    risk_keywords = [
        ("风险", "high"), ("问题", "medium"), ("困难", "medium"),
        ("延期", "high"), ("不足", "medium"), ("压力", "high"),
        ("投诉", "high"), ("竞品", "high"),
    ]
    lines = text.split("\n")
    for line in lines:
        line = line.strip()
        if not line:
            continue
        for kw, sev in risk_keywords:
            if kw in line:
                risks.append({
                    "customer": None,
                    "risk_description": line[:120],
                    "severity": sev,
                    "is_new": 1,
                    "source_column": "keyword_fallback",
                })
                break
    return risks[:20]


def run_risk_extraction(file_id, submission_file_id, module_id, user_id, week_start):
    """
    对一个文件执行风险提取并写入 DB。
    由 Worker 调用。
    """
    conn = get_db()
    sf = conn.execute("SELECT * FROM submission_files WHERE id = ?", (file_id,)).fetchone()
    if not sf:
        conn.close()
        return

    tmpl = conn.execute(
        "SELECT ft.name FROM file_templates ft WHERE ft.id = ?",
        (sf["template_id"],)
    ).fetchone()
    template_name = tmpl["name"] if tmpl else None
    conn.close()

    file_type = sf["file_type"]
    file_path = sf["original_path"]
    risks = []

    if file_type == "xlsx" and template_name:
        risks = extract_risks_from_excel(
            file_path, template_name, user_id, module_id, week_start, submission_file_id
        )
    elif file_type in ("pptx", "docx", "pdf"):
        risks = extract_risks_via_claude(file_path, file_type, module_id, week_start)

    if risks:
        upsert_risk_items(week_start, module_id, user_id, submission_file_id, risks)
