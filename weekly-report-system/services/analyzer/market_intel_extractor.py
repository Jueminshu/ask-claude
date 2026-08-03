"""
市场情报提取器
从销售部周报 Excel 中提取"市场信息（每周反馈新增变化部分）"区域
"""
import openpyxl


def extract_market_intel(file_path, module_id):
    """
    从 Excel 文件中提取市场情报数据。

    定位策略：遍历所有 sheet，在 column A 搜索含"市场信息"的单元格，
    下一行为表头，再下一行起为数据行。

    Args:
        file_path: Excel 文件路径
        module_id: 模块 ID（用于过滤，仅销售部=3）

    Returns:
        list[dict]: 每条记录的字段字典，提取失败返回 []
    """
    if module_id != 3:  # 仅销售部
        return []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return []

    all_rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 1. 定位"市场信息"section
        header_row = None

        for row_idx in range(1, ws.max_row + 1):
            cell_val = str(ws.cell(row=row_idx, column=1).value or "")
            if "市场信息" in cell_val:
                header_row = row_idx
                break

        if header_row is None:
            continue

        # 2. 表头行 = header_row + 1, 数据起始行 = header_row + 2
        if header_row + 1 > ws.max_row:
            continue

        # 3. 读取数据行
        data_start = header_row + 2
        for row_idx in range(data_start, ws.max_row + 1):
            row_vals = []
            for col_idx in range(1, 15):  # 14 列
                v = ws.cell(row=row_idx, column=col_idx).value
                row_vals.append(str(v).strip() if v is not None else None)

            # 判断空行：前3列全空则停止
            if not any(row_vals[:3]):
                break

            record = {
                "seq": row_vals[0],
                "update_time": row_vals[1],
                "collector": row_vals[2],
                "vendor": row_vals[3],
                "category": row_vals[4],
                "model": row_vals[5],
                "config": row_vals[6],
                "peripheral": row_vals[7],
                "price_tier": row_vals[8],
                "our_model": row_vals[9],
                "our_config": row_vals[10],
                "our_peripheral": row_vals[11],
                "our_price_tier": row_vals[12],
                "notes": row_vals[13],
            }
            all_rows.append(record)

    wb.close()
    return all_rows
