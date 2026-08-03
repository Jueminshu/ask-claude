"""
支持事项提取器
- Excel 模板（国内运营商、营销运营部、销售部）：从支持列直接读取
- PPT 模板（海外 BD）：Claude API 自由文本识别
"""
import os
import json
from database_v2 import get_db, upsert_support_items


# === 模板 → 支持列名映射 ===
TEMPLATE_SUPPORT_COLUMNS = {
    "周报": ["风险及求助"],            # 国内运营商
    "周报_营销": ["问题反馈&所需支持"],  # 营销运营部
}

# 支持相关关键词（用于模糊匹配列头）
SUPPORT_COLUMN_KEYWORDS = ["风险及求助", "问题反馈", "所需支持", "需要BU支持", "客户诉求"]


def _find_support_column(headers, template_name):
    """根据模板名查找支持列索引"""
    candidates = TEMPLATE_SUPPORT_COLUMNS.get(template_name, [])
    if not candidates:
        # 模糊匹配：任意 header 包含支持相关关键词
        for i, h in enumerate(headers):
            h_str = str(h).lower() if h else ""
            if any(kw.lower() in h_str for kw in SUPPORT_COLUMN_KEYWORDS):
                return i, str(h)
        return None, None
    for i, h in enumerate(headers):
        if str(h) in candidates:
            return i, str(h)
    return None, None


def _parse_excel_headers(file_path):
    """读取 Excel 文件所有数据 sheet 的 headers 和 rows。跳过目录/分析 sheet。"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        all_headers = []
        all_data_rows = []
        skip_keywords = ["目录", "分析", "汇总", "概述"]

        for name in wb.sheetnames:
            if any(kw in name for kw in skip_keywords):
                continue
            ws = wb[name]
            if ws.max_row < 2:
                continue
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and any(c for c in row):
                    all_data_rows.append([str(c) if c else "" for c in row])
            if not all_headers:
                all_headers = headers

        wb.close()
        return all_headers, all_data_rows
    except Exception:
        return [], []


def _extract_support_from_sales_excel(file_path):
    """
    销售部（module_id=3）专用提取：
    遍历 section，匹配列头含"需要BU支持"关键字，提取数据。
    通过 config.yaml 中定义的 start_row/end_row 定位各 section。
    """
    # 销售部模板 section 配置（与 config.yaml module_columns.sales.sections 一致）
    SALES_SECTIONS = [
        {
            "name": "日常工作",
            "start_row": 2,
            "end_row": 6,
            "num_columns": 5,
            "support_keywords": ["需要BU支持", "所需支持"],
        },
        {
            "name": "质量供应问题",
            "start_row": 11,
            "end_row": 15,
            "num_columns": 6,
            "support_keywords": ["客户诉求", "所需支持"],
        },
    ]

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return []

    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for section in SALES_SECTIONS:
            if section["start_row"] > ws.max_row:
                continue

            # 读取 section 表头行
            header_row = section["start_row"]
            headers = []
            for col_idx in range(1, section["num_columns"] + 1):
                val = ws.cell(row=header_row, column=col_idx).value
                headers.append(str(val) if val else "")

            # 查找支持列
            support_col_idx = None
            support_col_name = None
            for i, h in enumerate(headers):
                h_lower = h.lower()
                if any(kw.lower() in h_lower for kw in section["support_keywords"]):
                    support_col_idx = i  # 0-indexed
                    support_col_name = h
                    break

            if support_col_idx is None:
                continue

            # 查找客户列（通常为 section 第一列）
            customer_col_idx = None
            if headers:
                first_header = headers[0]
                if any(kw in first_header for kw in ["客户", "重点项目", "业务模块"]):
                    customer_col_idx = 0

            # 提取数据行
            for row_idx in range(section["start_row"] + 1, section["end_row"] + 1):
                if row_idx > ws.max_row:
                    break

                support_val = ws.cell(row=row_idx, column=support_col_idx + 1).value
                if not support_val or not str(support_val).strip():
                    continue

                customer = None
                if customer_col_idx is not None:
                    cust_val = ws.cell(row=row_idx, column=customer_col_idx + 1).value
                    customer = str(cust_val).strip() if cust_val else None

                results.append({
                    "customer": customer,
                    "support_description": str(support_val).strip(),
                    "source_column": f"{section['name']}/{support_col_name}",
                })

    wb.close()
    return results


def extract_support_from_excel(file_path, template_name, module_id, user_id, week_start, submission_file_id):
    """
    从 Excel 文件的支持列提取支持事项。
    返回: list[dict]
    """
    # 销售部（module_id=3）使用 section-based 提取
    if module_id == 3:
        return _extract_support_from_sales_excel(file_path)

    # 模块 1、2：标准列提取
    headers, data_rows = _parse_excel_headers(file_path)
    if not headers:
        return []

    support_col_idx, col_name = _find_support_column(headers, template_name)
    if support_col_idx is None:
        return []

    # 找到 customer 列（"重点项目"/"业务模块"/"客户"）
    customer_col_idx = None
    for i, h in enumerate(headers):
        h_str = str(h) if h else ""
        if any(kw in h_str for kw in ["重点项目", "业务模块", "客户"]):
            customer_col_idx = i
            break

    items = []
    for row in data_rows:
        if support_col_idx >= len(row) or not row[support_col_idx].strip():
            continue  # 支持列为空，跳过

        customer = row[customer_col_idx] if customer_col_idx is not None and customer_col_idx < len(row) else None
        support_text = row[support_col_idx].strip()

        items.append({
            "customer": customer,
            "support_description": support_text,
            "source_column": col_name,
        })

    return items


def extract_support_via_claude(file_path, file_type, module_id, week_start):
    """
    对 PPT 等非结构化文件，使用 Claude API 从文本中提取所需支持。
    返回: list[dict]
    """
    from services.file_parser import extract_text

    text = extract_text(file_path, file_type)
    if not text or text.startswith("["):
        return []  # 提取失败，返回空

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return _keyword_support_fallback(text)

    import urllib.request
    import urllib.error

    prompt = f"""你是一个周报分析助手。从以下周报文本中提取所有"所需支持"或"求助"条目。

周报文本：
{text[:6000]}

对每个所需支持条目，返回以下字段：
- customer: 关联的客户或项目名称（如无则填 null）
- support_description: 所需支持的具体描述（简要一句话）

请以 JSON 数组格式返回，只返回 JSON，不要其他文字。
如果无所需支持，返回空数组 []。

示例返回格式：
[{{"customer": "政企客户", "support_description": "需要BU提供产品技术方案支持"}}]"""

    try:
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=json.dumps({
                "model": "claude-sonnet-5",
                "max_tokens": 1024,
                "messages": [{"role": "user", "content": prompt}],
            }).encode("utf-8"),
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
            },
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read())
        content = result["content"][0]["text"]
        json_start = content.find("[")
        json_end = content.rfind("]") + 1
        if json_start >= 0 and json_end > json_start:
            items = json.loads(content[json_start:json_end])
            for item in items:
                item.setdefault("source_column", "claude_api")
            return items
        return []
    except Exception:
        return _keyword_support_fallback(text)


def _keyword_support_fallback(text):
    """关键词回退：从文本中识别支持请求信号"""
    items = []
    support_keywords = [
        ("需支持", True), ("求助", True), ("需要BU", True),
        ("协调", True), ("资源不足", True), ("协助", False),
    ]
    lines = text.split("\n")
    for line in lines:
        line = line.strip()
        if not line:
            continue
        for kw, _ in support_keywords:
            if kw in line:
                items.append({
                    "customer": None,
                    "support_description": line[:120],
                    "source_column": "keyword_fallback",
                })
                break
    return items[:20]


def run_support_extraction(file_id, submission_file_id, module_id, user_id, week_start):
    """
    对一个文件执行支持事项提取并写入 DB。
    由 Worker 调用。
    """
    conn = get_db()
    sf = conn.execute("SELECT * FROM submission_files WHERE id = ?", (file_id,)).fetchone()
    if not sf:
        conn.close()
        return

    tmpl = conn.execute(
        "SELECT ft.name FROM file_templates ft WHERE ft.id = ?",
        (sf["template_id"],)
    ).fetchone()
    template_name = tmpl["name"] if tmpl else None
    conn.close()

    file_type = sf["file_type"]
    file_path = sf["original_path"]
    items = []

    if file_type == "xlsx" and template_name:
        items = extract_support_from_excel(
            file_path, template_name, module_id, user_id, week_start, submission_file_id
        )
    elif file_type in ("pptx", "docx", "pdf"):
        items = extract_support_via_claude(file_path, file_type, module_id, week_start)

    # 过滤空字段：support_description 为空的不列入
    items = [s for s in items if s.get("support_description") and str(s["support_description"]).strip()]
    if items:
        upsert_support_items(week_start, module_id, user_id, submission_file_id, items)
